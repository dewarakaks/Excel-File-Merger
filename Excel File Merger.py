# Mengimpor modul yang diperlukan
import pandas as pd  # Untuk mengelola data dalam format DataFrame
import os  # Untuk berinteraksi dengan sistem file dan folder
import openpyxl  # Untuk membaca dan memodifikasi file Excel

# Mendapatkan path direktori saat ini
path = os.getcwd()  # Mengambil path direktori kerja saat ini

# Mengambil daftar file yang ada di dalam folder
files = os.listdir(path)  # Daftar semua file di direktori saat ini

# Menyaring hanya file dengan ekstensi '.xlsx'
all_files = [f for f in files if f[-4:] == 'xlsx']  # Filter file Excel (.xlsx)

# Inisialisasi list untuk menyimpan data dari setiap file
li = []  # List untuk menyimpan DataFrame yang digabungkan

# Loop untuk membaca setiap file Excel
for filename in all_files:
    # Membaca file Excel ke dalam DataFrame menggunakan pandas
    df = pd.read_excel(filename, index_col=None, header=0)  # Membaca file Excel tanpa menetapkan kolom indeks

    # Menambahkan kolom 'Brand' ke DataFrame, diisi dengan nama brand yang didapatkan dari nama file
    # Nama brand diambil dari posisi 12 hingga karakter pertama setelah tanda '-' dalam nama file
    print(filename[12:(filename.find("-")-1)])  # Menampilkan nama brand yang diambil dari nama file
    df["Brand"] = filename[12:(filename.find("-")-1)]  # Menambahkan kolom 'Brand'
    
    # Menampilkan nama file
    print(filename)

    # Membuka file Excel menggunakan openpyxl untuk mengakses worksheet dan hyperlink
    wb = openpyxl.load_workbook(filename, data_only=True)  # Membuka file Excel, hanya data (bukan formula)
    
    # Mengakses sheet bernama 'Data'
    ws = wb['Data']  # Mengambil sheet bernama 'Data'
    
    # Inisialisasi list untuk menyimpan hyperlink dari kolom 'A'
    link = []

    # Loop untuk membaca semua hyperlink yang ada di kolom 'A' pada sheet 'Data'
    for x in range(len(ws["A"]) - 1):  # Mengiterasi setiap baris di kolom A (baris pertama adalah header)
        link.append(ws.cell(row=x + 2, column=1).hyperlink.target)  # Menyimpan target hyperlink pada baris x+2 (baris pertama adalah header)

    # Menambahkan kolom 'Link' ke DataFrame dengan hyperlink yang diambil
    df["Link"] = link

    # Menambahkan DataFrame ini ke list li
    li.append(df)

# Menggabungkan semua DataFrame dari list 'li' menjadi satu DataFrame
data = pd.concat(li, axis=0, ignore_index=True)  # Menggabungkan DataFrame secara vertikal (baris per baris)

# Mengisi nilai yang kosong (NaN) dengan 0
data = data.fillna(0)  # Mengganti NaN dengan 0

# Menambahkan kolom 'Date' dengan tanggal yang diinginkan (dalam hal ini 11/1/2024)
data['Date'] = "11/1/2024"  # Menambahkan kolom 'Date' dengan nilai tanggal statis

# Menyimpan DataFrame yang sudah digabungkan dan diproses ke file Excel baru
data.to_excel("Tokopedia_11-24.xlsx")  # Menyimpan DataFrame ke file Excel dengan nama 'Tokopedia_11-24.xlsx'

# Bagian berikutnya adalah kode yang dikomentari, mungkin untuk menggunakan Google Sheets API, tetapi tidak aktif
# gc = gspread.oauth(
#     credentials_filename='C:/Users/data analyst/Documents/Python Project/Brand/client_secret.json'
# )

# sh = gc.open("Trial")  # Membuka spreadsheet 'Trial' di Google Sheets

# #print(sh.sheet1.get('A1'))  # Menampilkan nilai dari sel A1
# worksheet = sh.get_worksheet(0)  # Mengakses worksheet pertama
# worksheet = sh.add_worksheet(title="Tokopedia", rows=10, cols=10)  # Menambahkan worksheet baru dengan judul 'Tokopedia'
# #print(data.values)  # Menampilkan data yang akan dimasukkan ke worksheet
# worksheet.update([data.columns.values.tolist()] + data.values.tolist())  # Memperbarui worksheet dengan data dari DataFrame
# #worksheet.update_cell(1, 2, 'Bingo!')  # Mengupdate sel tertentu dengan nilai 'Bingo!'
